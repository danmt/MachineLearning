from openpyxl import load_workbook
from arquitecturas.RBF.RedNeuronal import RedNeuronal as RBF
from arquitecturas.SOM_1D.RedNeuronal import RedNeuronal as SOM_1D
import config_20_centros
import config_10_centros


def leer_datos(file):
	wb2 = load_workbook(file)
		
	hoja = wb2.get_active_sheet()

	celda = hoja.cell(row=1,column=1)

	x = []
	y = []

	for i in xrange(1,20):
		xi = hoja.cell(row=i,column=1).value
		yi = hoja.cell(row=i,column=2).value
		x.append(xi)
		y.append(yi)

	return [x,y]

def generar_centros_SOM():
	som = SOM_1D(config_10_centros.tam_entrada(),config_10_centros.tam_oculta())

	centros = config_10_centros.centros()

	tasa_aprendizaje = 0.1
	epocas = 10000

	som.entrenar(centros,tasa_aprendizaje,epocas)

	return som.obtener_centros()

def generar_centros_manualmente():
	centros = config_10_centros.centros()

	return centros

def generar_red_RBF(estrategia):
	#Funciones de activacion
	funcion_oculta = "gauss"
	funcion_salida = "ident"

	#Datos
	if (estrategia == 'manualmente'):
		centros = generar_centros_manualmente()
	elif (estrategia == 'SOM'):
		centros = generar_centros_SOM()
	else:
		print("Estrategia desconocida.")
		return False

	anchuras = config_10_centros.anchuras()
	pesos = config_10_centros.pesos()
	sesgos = config_10_centros.sesgos()

	rbf = RBF(config_10_centros.tam_entrada(),config_10_centros.tam_oculta(),config_10_centros.tam_salida(),funcion_oculta,funcion_salida)	
	
	#Se asignan los datos correspondientes 
	rbf.asignar_pesos(pesos)
	rbf.asignar_centros(centros)
	rbf.asignar_sesgos(sesgos)
	rbf.asignar_anchuras(anchuras)

	return rbf

def graficar_RBF(estrategia,entrada,datos,plot):
	rbf = generar_red_RBF(estrategia)

	if not rbf:
		return False

	#Se obtienen los datos del problema
	x = datos[0]
	y = datos[1]

	#SE GRAFICAN LA ENTRADA
	plot.set_title('Escogencia de centros: ' + estrategia)
	plot.scatter(x, y)

	#SE GRAFICA LA SALIDA
	rbf.graficar_salida(entrada,plot)

	#SE GRAFICAN LOS CENTROS
	rbf.graficar_centros(plot)

	#SE GRAFICAN LAS FUNCIONES
	rbf.graficar_funciones(entrada,plot)